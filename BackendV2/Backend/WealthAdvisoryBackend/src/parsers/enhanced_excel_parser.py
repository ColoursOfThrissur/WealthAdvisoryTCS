"""
Enhanced Excel Parser for Professional Reports
Parses 6-sheet Excel format with complete client data
"""
import openpyxl
import pandas as pd
from pathlib import Path
from typing import Dict, List
from datetime import datetime


class EnhancedExcelParser:
    """Parse professional multi-sheet Excel portfolio files"""
    
    def parse(self, file_path: str) -> Dict:
        """
        Parse Excel file with up to 6 sheets:
        1. Client_Info: firm, advisor, client details
        2. Accounts: account names, types, tax status
        3. Holdings: ticker, shares, cost basis, asset class
        4. Target_Allocation: asset class targets
        5. Transactions: full transaction history
        6. Goals: retirement, education goals (optional)
        
        Returns complete portfolio data structure
        """
        path = Path(file_path)
        
        if not path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")
        
        wb = openpyxl.load_workbook(file_path, data_only=True)
        
        result = {
            "client_info": self._parse_client_info(wb),
            "accounts": self._parse_accounts(wb),
            "holdings": self._parse_holdings(wb),
            "target_allocation": self._parse_target_allocation(wb),
            "transactions": self._parse_transactions(wb),
            "goals": self._parse_goals(wb)
        }
        
        return result
    
    def _parse_client_info(self, wb) -> Dict:
        """Parse Client_Info sheet"""
        if "Client_Info" not in wb.sheetnames:
            return {}
        
        ws = wb["Client_Info"]
        info = {}
        
        headers = [cell.value for cell in ws[1]]
        for row in ws.iter_rows(min_row=2, max_row=2, values_only=True):
            for i, header in enumerate(headers):
                if header and i < len(row) and row[i]:
                    key = str(header).lower().replace(" ", "_")
                    info[key] = row[i]
            break
        
        return info
    
    def _parse_accounts(self, wb) -> List[Dict]:
        """Parse Accounts sheet"""
        if "Accounts" not in wb.sheetnames:
            return []
        
        ws = wb["Accounts"]
        accounts = []
        
        headers = [str(cell.value).lower().replace(" ", "_") if cell.value else None 
                  for cell in ws[1]]
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            
            account = {}
            for i, header in enumerate(headers):
                if header and i < len(row) and row[i]:
                    account[header] = row[i]
            
            if account:
                accounts.append(account)
        
        return accounts
    
    def _parse_holdings(self, wb) -> List[Dict]:
        """Parse Holdings sheet with asset class"""
        if "Holdings" not in wb.sheetnames:
            ws = wb.active
        else:
            ws = wb["Holdings"]
        
        holdings = []
        headers = [str(cell.value).lower().replace(" ", "_") if cell.value else None 
                  for cell in ws[1]]
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            
            holding = {}
            for i, header in enumerate(headers):
                if header and i < len(row) and row[i] is not None:
                    if header in ['ticker', 'symbol']:
                        holding['ticker'] = str(row[i]).strip().upper()
                    elif header in ['security', 'security_name', 'name']:
                        holding['security'] = str(row[i])
                    elif header in ['shares', 'quantity']:
                        holding['shares'] = float(row[i])
                    elif header in ['cost_basis', 'cost', 'basis']:
                        holding['cost_basis'] = float(row[i])
                    elif header in ['account', 'account_name']:
                        holding['account'] = str(row[i])
                    elif header in ['asset_class', 'class', 'type']:
                        holding['asset_class'] = str(row[i])
            
            if 'ticker' in holding and 'shares' in holding:
                holdings.append(holding)
        
        return holdings
    
    def _parse_target_allocation(self, wb) -> Dict:
        """Parse Target_Allocation sheet"""
        if "Target_Allocation" not in wb.sheetnames:
            return {}
        
        ws = wb["Target_Allocation"]
        targets = {}
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            
            asset_class = str(row[0])
            target_pct = float(row[1]) if len(row) > 1 and row[1] else 0.0
            
            targets[asset_class] = target_pct
        
        return targets
    
    def _parse_transactions(self, wb) -> List[Dict]:
        """Parse Transactions sheet"""
        if "Transactions" not in wb.sheetnames:
            return []
        
        ws = wb["Transactions"]
        transactions = []
        
        headers = [str(cell.value).lower().replace(" ", "_") if cell.value else None 
                  for cell in ws[1]]
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            
            txn = {}
            for i, header in enumerate(headers):
                if header and i < len(row) and row[i] is not None:
                    if header in ['date', 'transaction_date', 'trade_date']:
                        if isinstance(row[i], datetime):
                            txn['date'] = row[i].strftime('%Y-%m-%d')
                        else:
                            txn['date'] = str(row[i])
                    elif header in ['type', 'transaction_type', 'txn_type']:
                        txn['type'] = str(row[i])
                    elif header in ['ticker', 'symbol']:
                        txn['ticker'] = str(row[i]).strip().upper()
                    elif header in ['description', 'desc', 'memo']:
                        txn['description'] = str(row[i])
                    elif header in ['shares', 'quantity']:
                        txn['shares'] = float(row[i])
                    elif header in ['amount', 'value', 'price']:
                        txn['amount'] = float(row[i])
                    elif header in ['account', 'account_name']:
                        txn['account'] = str(row[i])
            
            if 'date' in txn and 'type' in txn:
                transactions.append(txn)
        
        return transactions
    
    def _parse_goals(self, wb) -> List[Dict]:
        """Parse Goals sheet (optional)"""
        if "Goals" not in wb.sheetnames:
            return []
        
        ws = wb["Goals"]
        goals = []
        
        headers = [str(cell.value).lower().replace(" ", "_") if cell.value else None 
                  for cell in ws[1]]
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            
            goal = {}
            for i, header in enumerate(headers):
                if header and i < len(row) and row[i] is not None:
                    goal[header] = row[i]
            
            if goal:
                goals.append(goal)
        
        return goals
